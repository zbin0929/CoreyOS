#!/usr/bin/env python3
import sys
import json
import os
import re
import struct
import subprocess
import tempfile
import argparse

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None


def _get_zone_index():
    import urllib.request
    url = "https://www.ups.com/us/en/zone-chart.json"
    req = urllib.request.Request(url, headers={
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
        "Accept": "application/json",
    })
    with urllib.request.urlopen(req, timeout=60) as resp:
        return json.loads(resp.read().decode())


def _download_xls(url, dest):
    import urllib.request
    full_url = f"https://assets.ups.com{url}" if url.startswith("/") else url
    req = urllib.request.Request(full_url, headers={
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
        "Accept": "*/*",
        "Referer": "https://www.ups.com/us/en/support/shipping-support/shipping-costs-rates/daily-rates",
    })
    with urllib.request.urlopen(req, timeout=120) as resp:
        data = resp.read()
        with open(dest, "wb") as f:
            f.write(data)
    return len(data)


def _download_via_curl(url, dest, cookie_file=None):
    full_url = f"https://assets.ups.com{url}" if url.startswith("/") else url
    cmd = [
        "curl", "-sL", "-o", dest,
        "-H", "User-Agent: Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
        "-H", "Accept: */*",
        "-H", "Referer: https://www.ups.com/us/en/support/shipping-support/shipping-costs-rates/daily-rates",
        "--connect-timeout", "30",
        "--max-time", "120",
        full_url,
    ]
    if cookie_file and os.path.exists(cookie_file):
        cmd.extend(["-b", cookie_file])
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=180)
    if result.returncode != 0:
        raise RuntimeError(f"curl failed: {result.stderr}")
    return os.path.getsize(dest)


def parse_xls_ground(xls_path):
    if xlrd is None:
        raise ImportError("xlrd is required: pip install xlrd==1.2.0")
    wb = xlrd.open_workbook(xls_path, formatting_info=False)
    sheet = wb.sheet_by_index(0)

    header_row = None
    ground_col = None
    zip_col = None
    for r in range(sheet.nrows):
        row_vals = [str(sheet.cell_value(r, c)).strip() for c in range(sheet.ncols)]
        for i, v in enumerate(row_vals):
            v_lower = v.lower()
            if "ground" in v_lower and "air" not in v_lower:
                header_row = r
                ground_col = i
            if ("zip" in v_lower or "dest" in v_lower) and zip_col is None:
                zip_col = i
        if header_row is not None:
            break

    if header_row is None or ground_col is None:
        raise RuntimeError(f"Cannot find Ground column header in {xls_path}")
    if zip_col is None:
        zip_col = 0

    main_rows = []
    footnotes = []
    for r in range(header_row + 1, sheet.nrows):
        dest_zip_raw = str(sheet.cell_value(r, zip_col)).strip()
        ground_raw = str(sheet.cell_value(r, ground_col)).strip()

        row0 = str(sheet.cell_value(r, 0)).strip()
        if (row0.startswith("[") or row0.lower().startswith("for ")) and "zone" in row0.lower():
            zone_match = re.search(r"Zone\s+(\d+)\s+for\s+Ground", row0, re.IGNORECASE)
            if zone_match:
                fn_zone = zone_match.group(1)
                fn_zips = []
                for rr in range(r + 1, sheet.nrows):
                    rr0 = str(sheet.cell_value(rr, 0)).strip()
                    if not rr0:
                        break
                    if (rr0.startswith("[") or rr0.lower().startswith("for ")) and "zone" in rr0.lower():
                        break
                    for c in range(sheet.ncols):
                        cv = str(sheet.cell_value(rr, c)).strip()
                        if cv and re.match(r"^\d{4,5}\.0$", cv):
                            zip5 = cv.replace(".0", "").zfill(5)
                            fn_zips.append(zip5)
                footnotes.append((fn_zone, fn_zips))
            continue

        if not dest_zip_raw or not ground_raw:
            continue
        if not re.match(r"^\d{3}$", dest_zip_raw):
            continue
        if ground_raw in ("-", "--", ""):
            continue

        ground = ground_raw.lstrip("0")
        if not ground:
            ground = "0"
        main_rows.append((dest_zip_raw, ground))

    footnote_rows = []
    for fn_zone, fn_zips in footnotes:
        for z5 in fn_zips:
            footnote_rows.append((z5, fn_zone))

    return main_rows, footnote_rows


def ground_to_zone_code(ground):
    if ground in ("-", "--"):
        return None
    if ground.startswith("["):
        ground = ground.strip("[]")
    zone_num = ground.lstrip("0")
    if not zone_num:
        return None
    return f"Zone{zone_num}"


def expand_zip_range(zip_str):
    zip_str = zip_str.strip()
    if "-" in zip_str:
        parts = zip_str.split("-")
        start3 = parts[0].strip()
        end3 = parts[1].strip()
        return (f"{start3}00", f"{end3}99")
    else:
        return (f"{zip_str}00", f"{zip_str}99")


def build_template_rows(main_rows, footnote_rows=None):
    template = []
    for dest_zip, ground in main_rows:
        zone_code = ground_to_zone_code(ground)
        if not zone_code:
            continue
        start_zip, end_zip = expand_zip_range(dest_zip)
        template.append((zone_code, start_zip, end_zip))
    if footnote_rows:
        for zip5, zone in footnote_rows:
            zone_code = f"Zone{zone}"
            template.append((zone_code, zip5, zip5))
    return template


def write_excel(template_rows, output_path):
    if openpyxl is None:
        raise ImportError("openpyxl is required: pip install openpyxl")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    notice = "1、分区代码、开始邮编、截止邮编均为必填\n2、邮编仅支持数字、字母\n3、仅能填写已添加的分区代码\n4、同一个邮编不能同时存在于2个分区内"
    ws.append([notice, "", ""])
    ws.append(["分区代码", "开始邮编", "截止邮编"])
    for row in template_rows:
        ws.append(list(row))
    for col_idx in range(2, 4):
        col_letter = get_column_letter(col_idx)
        for row_idx in range(3, len(template_rows) + 3):
            cell = ws[f"{col_letter}{row_idx}"]
            cell.number_format = "@"
    wb.save(output_path)


def download_and_convert(zip3, output_path, cookie_file=None):
    print(f"Fetching zone chart index...", file=sys.stderr)
    index = _get_zone_index()
    print(f"Index loaded: {index['total']} ZIP3 entries", file=sys.stderr)

    entry = None
    for d in index["data"]:
        if d["zip"] == zip3:
            entry = d
            break
    if not entry:
        raise RuntimeError(f"ZIP3 {zip3} not found in index")

    xls_url = entry["url"]
    print(f"Found XLS URL for {zip3}: {xls_url}", file=sys.stderr)

    with tempfile.NamedTemporaryFile(suffix=".xls", delete=False) as tmp:
        tmp_xls = tmp.name

    try:
        print(f"Downloading XLS...", file=sys.stderr)
        try:
            size = _download_xls(xls_url, tmp_xls)
        except Exception as e:
            print(f"urllib failed ({e}), trying curl...", file=sys.stderr)
            size = _download_via_curl(xls_url, tmp_xls, cookie_file)
        print(f"Downloaded: {size} bytes", file=sys.stderr)

        if size < 100:
            raise RuntimeError(f"Downloaded file too small ({size} bytes), likely blocked")

        print(f"Parsing XLS...", file=sys.stderr)
        main_rows, footnote_rows = parse_xls_ground(tmp_xls)
        print(f"Parsed {len(main_rows)} main rows, {len(footnote_rows)} footnote rows", file=sys.stderr)

        template = build_template_rows(main_rows, footnote_rows)
        print(f"Built {len(template)} template rows", file=sys.stderr)

        write_excel(template, output_path)
        print(f"Written to {output_path}", file=sys.stderr)
        return output_path, len(template)
    finally:
        if os.path.exists(tmp_xls):
            os.unlink(tmp_xls)


def load_checkpoint(path):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"completed": [], "failed": []}


def save_checkpoint(path, data):
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


def batch_download(output_dir, checkpoint_path=None):
    import time
    if checkpoint_path is None:
        checkpoint_path = os.path.join(output_dir, "checkpoint.json")

    os.makedirs(output_dir, exist_ok=True)
    cp = load_checkpoint(checkpoint_path)
    completed = set(cp.get("completed", []))

    print(f"Fetching zone index...", file=sys.stderr)
    index = _get_zone_index()
    total = len(index["data"])
    print(f"Index: {total} entries, {len(completed)} already done", file=sys.stderr)

    new_failed = []
    for entry in index["data"]:
        zip3 = entry["zip"]
        if zip3 in completed:
            continue

        xlsx_path = os.path.join(output_dir, f"UPS-GROUND-{zip3}.xlsx")
        done = len(completed) + len(new_failed) + 1
        print(f"[{done}/{total}] {zip3}...", end=" ", file=sys.stderr)

        try:
            _, row_count = download_and_convert(zip3, xlsx_path)
            completed.add(zip3)
            cp["completed"] = sorted(completed)
            save_checkpoint(checkpoint_path, cp)
            print(f"OK ({row_count})", file=sys.stderr)
        except Exception as e:
            new_failed.append({"zip3": zip3, "error": str(e)})
            print(f"FAIL: {e}", file=sys.stderr)

        time.sleep(0.5)

    if new_failed:
        cp["failed"] = cp.get("failed", []) + new_failed
        save_checkpoint(checkpoint_path, cp)

    print(f"\nDone: {len(completed)} OK, {len(new_failed)} failed", file=sys.stderr)
    return {"total": total, "completed": len(completed), "failed": len(new_failed)}


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Download UPS zone XLS and convert to template")
    parser.add_argument("--zip3", default="", help="3-digit ZIP code (e.g. 910)")
    parser.add_argument("--all", action="store_true", help="Download all 902 ZIP3 prefixes")
    parser.add_argument("--output", default="", help="Output xlsx path (single) or directory (batch)")
    parser.add_argument("--cookie-file", default="", help="Cookie file for curl (optional)")
    args = parser.parse_args()

    if not args.zip3 and not args.all:
        parser.error("Specify --zip3 <code> or --all")

    try:
        if args.zip3:
            output = args.output or os.path.expanduser(f"~/Desktop/UPS-GROUND-{args.zip3}.xlsx")
            result_path, row_count = download_and_convert(args.zip3, output, args.cookie_file or None)
            print(json.dumps({
                "status": "success",
                "output": result_path,
                "rows": row_count,
            }, ensure_ascii=False))
        elif args.all:
            output_dir = args.output or os.path.join(
                os.path.expanduser("~"), ".hermes", "pack-data", "meizheng", "zone-charts"
            )
            result = batch_download(output_dir)
            print(json.dumps(result, ensure_ascii=False))
    except Exception as e:
        print(json.dumps({"status": "failed", "reason": str(e)}, ensure_ascii=False))
        sys.exit(1)
