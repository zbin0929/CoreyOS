#!/usr/bin/env python3
import sys, os, json, time, re, tempfile, argparse
import requests
from datetime import datetime

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

BASE_URL = "https://postcalc.usps.com/DomesticZoneChart"
API_URL = "https://postcalc.usps.com/DomesticZoneChart/GetZoneChart"
ZONE_NAME_PREFIX = "USPS-GROUND"
HERMES_DIR = os.environ.get("HERMES_DIR", os.path.expanduser("~/.hermes"))
CONFIG_PATH = os.path.join(HERMES_DIR, "pack-data", "meizheng", "config", "zone-config.yaml")
TOKEN_CACHE_PATH = os.path.join(HERMES_DIR, "pack-data", "meizheng", "config", ".usps_token_cache.json")


def create_session():
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    })
    r = session.get(BASE_URL, timeout=30)
    if r.status_code != 200:
        print(f"Failed to init session: {r.status_code}", file=sys.stderr)
        return None
    return session


def fetch_zone_data(zip3, session):
    zip3 = str(zip3).zfill(3)
    today = datetime.now().strftime("%m/%d/%Y")
    resp = session.get(API_URL, params={
        "zipCode3Digit": zip3,
        "shippingDate": today,
    }, timeout=30)

    data = resp.json()
    zip_err = data.get("ZIPCodeError", "")
    if zip_err:
        print(f"ZIP error for {zip3}: {zip_err}", file=sys.stderr)
        return None
    page_err = data.get("PageError", "")
    if page_err and "No Zones" in page_err:
        print(f"No zones for {zip3}", file=sys.stderr)
        return None

    all_cols = []
    for i in range(4):
        all_cols.extend(data.get(f"Column{i}", []))

    zip5 = data.get("Zip5Digit", [])
    effective = data.get("EffectiveDate", "")

    zones = []
    for item in all_cols:
        raw_zip = item.get("ZipCodes", "").strip()
        raw_zone = item.get("Zone", "").strip()
        mail_svc = item.get("MailService", "")
        if not raw_zip or not raw_zone:
            continue
        zone_num = re.sub(r"[^0-9]", "", raw_zone)
        has_plus = "+" in raw_zone
        if "---" in raw_zip:
            parts = raw_zip.split("---")
            dest_start = parts[0].strip().zfill(3)
            dest_end = parts[1].strip().zfill(3)
        else:
            dest_start = raw_zip.zfill(3)
            dest_end = raw_zip.zfill(3)
        zones.append({
            "zone": zone_num,
            "dest_start": dest_start,
            "dest_end": dest_end,
            "has_5digit_exceptions": has_plus,
            "priority_mail_only": mail_svc == "Priority Mail",
        })

    for exc in zip5:
        raw_zip = str(exc.get("ZipCodes", "")).strip()
        raw_zone = str(exc.get("Zone", "")).strip()
        zone_num = re.sub(r"[^0-9]", "", raw_zone)
        if not raw_zip or not zone_num:
            continue
        if "---" in raw_zip:
            parts = raw_zip.split("---")
            dest_start = parts[0].strip()
            dest_end = parts[1].strip()
        else:
            dest_start = raw_zip
            dest_end = raw_zip
        zones.append({
            "zone": zone_num,
            "dest_start": dest_start,
            "dest_end": dest_end,
            "has_5digit_exceptions": False,
            "priority_mail_only": False,
        })

    return {"zip3": zip3, "effective": effective, "zones": zones}


def build_excel_rows(zone_data):
    rows = []
    for z in zone_data["zones"]:
        zone_code = f"Zone{z['zone']}"
        start = z["dest_start"]
        end = z["dest_end"]
        if len(start) == 3:
            start = start + "00"
            end = end + "99"
        rows.append((zone_code, start, end))
    return rows


def write_excel(rows, output_path):
    if openpyxl is None:
        raise ImportError("openpyxl is required: pip install openpyxl")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    notice = "1、分区代码、开始邮编、截止邮编均为必填\n2、邮编仅支持数字、字母\n3、仅能填写已添加的分区代码\n4、同一个邮编不能同时存在于2个分区内"
    ws.append([notice, "", ""])
    ws.append(["分区代码", "开始邮编", "截止邮编"])
    for row in rows:
        ws.append(list(row))
    for col_idx in range(2, 4):
        col_letter = get_column_letter(col_idx)
        for row_idx in range(3, len(rows) + 3):
            ws[f"{col_letter}{row_idx}"].number_format = "@"
    wb.save(output_path)


def download_and_convert(zip3, output_dir, session=None):
    if session is None:
        session = create_session()
        if session is None:
            return None

    zone_data = fetch_zone_data(zip3, session)
    if not zone_data:
        return None

    print(f"  {len(zone_data['zones'])} zone entries, effective {zone_data['effective']}", file=sys.stderr)

    rows = build_excel_rows(zone_data)
    if not rows:
        print(f"  No rows to write for {zip3}", file=sys.stderr)
        return None

    os.makedirs(output_dir, exist_ok=True)
    xlsx_name = f"USPS-GROUND-{str(zip3).zfill(3)}.xlsx"
    xlsx_path = os.path.join(output_dir, xlsx_name)
    write_excel(rows, xlsx_path)
    print(f"  Saved: {xlsx_path} ({len(rows)} rows)", file=sys.stderr)
    return xlsx_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Download USPS zone charts and convert to template Excel")
    parser.add_argument("--zip3", default=None, help="Single ZIP3 code")
    parser.add_argument("--zip3-list", default=None, help="File with ZIP3 codes, one per line")
    parser.add_argument("--output-dir", default=os.path.join(os.path.expanduser("~"), "Desktop"))
    args = parser.parse_args()

    if args.zip3:
        result = download_and_convert(args.zip3, args.output_dir)
        if result:
            print(f"SUCCESS: {result}")
        else:
            print("FAILED")
            sys.exit(1)
    elif args.zip3_list:
        session = create_session()
        if not session:
            print("FAILED to create session")
            sys.exit(1)
        with open(args.zip3_list) as f:
            zip3_list = [line.strip() for line in f if line.strip()]
        results = []
        for i, zip3 in enumerate(zip3_list):
            print(f"\n[{i+1}/{len(zip3_list)}] ZIP3 {zip3}", file=sys.stderr)
            r = download_and_convert(zip3, args.output_dir, session=session)
            if r:
                results.append(r)
            time.sleep(0.5)
        print(f"\nDone: {len(results)}/{len(zip3_list)} succeeded")
    else:
        print("Specify --zip3 or --zip3-list", file=sys.stderr)
        sys.exit(1)
