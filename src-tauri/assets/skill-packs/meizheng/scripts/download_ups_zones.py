#!/usr/bin/env python3
import sys
import json
import os
import re
import time
import tempfile
import argparse
import urllib.request

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None

INDEX_URL = "https://www.ups.com/us/en/zone-chart.json"
ASSET_BASE = "https://assets.ups.com"
REFERER = "https://www.ups.com/us/en/support/shipping-support/shipping-costs-rates/daily-rates"
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
MAX_RETRIES = 3
RETRY_DELAYS = [2, 4, 8]
DOWNLOAD_TIMEOUT = 120


def _http_get_json(url, timeout=60):
    req = urllib.request.Request(url, headers={
        "User-Agent": UA,
        "Accept": "application/json",
    })
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode())


def _http_download(url, dest):
    req = urllib.request.Request(url, headers={
        "User-Agent": UA,
        "Accept": "*/*",
        "Referer": REFERER,
    })
    with urllib.request.urlopen(req, timeout=DOWNLOAD_TIMEOUT) as resp:
        data = resp.read()
        with open(dest, "wb") as f:
            f.write(data)
    return len(data)


def get_zone_index():
    return _http_get_json(INDEX_URL)


def find_xls_url(index, zip3):
    for d in index["data"]:
        if d["zip"] == zip3:
            return d["url"]
    return None


def download_xls(xls_path_url, dest):
    full_url = f"{ASSET_BASE}{xls_path_url}" if xls_path_url.startswith("/") else xls_path_url
    return _http_download(full_url, dest)


def parse_xls_ground(xls_path):
    if xlrd is None:
        raise ImportError("xlrd is required: pip install xlrd==1.2.0")
    wb = xlrd.open_workbook(xls_path, formatting_info=False)
    sheet = wb.sheet_by_index(0)

    header_row = None
    ground_col = None
    zip_col = None
    for r in range(sheet.nrows):
        row_vals = [str(sheet.cell_value(r, c)).strip() for c in range(sheet.ncols)]
        for i, v in enumerate(row_vals):
            vl = v.lower()
            if "ground" in vl and "air" not in vl:
                header_row = r
                ground_col = i
            if ("zip" in vl or "dest" in vl) and zip_col is None:
                zip_col = i
        if header_row is not None:
            break

    if header_row is None or ground_col is None:
        raise RuntimeError(f"Cannot find Ground column header")
    if zip_col is None:
        zip_col = 0

    main_rows = []
    footnotes = []
    for r in range(header_row + 1, sheet.nrows):
        dest_zip_raw = str(sheet.cell_value(r, zip_col)).strip()
        ground_raw = str(sheet.cell_value(r, ground_col)).strip()
        row0 = str(sheet.cell_value(r, 0)).strip()

        if (row0.startswith("[") or row0.lower().startswith("for ")) and "zone" in row0.lower():
            zone_match = re.search(r"Zone\s+(\d+)\s+for\s+Ground", row0, re.IGNORECASE)
            if zone_match:
                fn_zone = zone_match.group(1)
                fn_zips = []
                for rr in range(r + 1, sheet.nrows):
                    rr0 = str(sheet.cell_value(rr, 0)).strip()
                    if not rr0:
                        break
                    if (rr0.startswith("[") or rr0.lower().startswith("for ")) and "zone" in rr0.lower():
                        break
                    for c in range(sheet.ncols):
                        cv = str(sheet.cell_value(rr, c)).strip()
                        if cv and re.match(r"^\d{4,5}\.0$", cv):
                            fn_zips.append(cv.replace(".0", "").zfill(5))
                footnotes.append((fn_zone, fn_zips))
            continue

        if not dest_zip_raw or not ground_raw:
            continue
        if not re.match(r"^\d{3}$", dest_zip_raw):
            continue
        if ground_raw in ("-", "--", ""):
            continue

        ground = ground_raw.lstrip("0")
        if not ground:
            ground = "0"
        main_rows.append((dest_zip_raw, ground))

    footnote_rows = []
    for fn_zone, fn_zips in footnotes:
        for z5 in fn_zips:
            footnote_rows.append((z5, fn_zone))

    return main_rows, footnote_rows


def ground_to_zone_code(ground):
    if ground in ("-", "--"):
        return None
    if ground.startswith("["):
        ground = ground.strip("[]")
    zone_num = ground.lstrip("0")
    if not zone_num:
        return None
    return f"Zone{zone_num}"


def expand_zip_range(zip_str):
    zip_str = zip_str.strip()
    if "-" in zip_str:
        parts = zip_str.split("-")
        return f"{parts[0].strip()}00", f"{parts[1].strip()}99"
    return f"{zip_str}00", f"{zip_str}99"


def build_template_rows(main_rows, footnote_rows=None):
    template = []
    for dest_zip, ground in main_rows:
        zone_code = ground_to_zone_code(ground)
        if not zone_code:
            continue
        start_zip, end_zip = expand_zip_range(dest_zip)
        template.append((zone_code, start_zip, end_zip))
    if footnote_rows:
        for zip5, zone in footnote_rows:
            template.append((f"Zone{zone}", zip5, zip5))
    return template


def write_excel(template_rows, output_path):
    if openpyxl is None:
        raise ImportError("openpyxl is required: pip install openpyxl")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    notice = "1、分区代码、开始邮编、截止邮编均为必填\n2、邮编仅支持数字、字母\n3、仅能填写已添加的分区代码\n4、同一个邮编不能同时存在于2个分区内"
    ws.append([notice, "", ""])
    ws.append(["分区代码", "开始邮编", "截止邮编"])
    for row in template_rows:
        ws.append(list(row))
    for col_idx in range(2, 4):
        col_letter = get_column_letter(col_idx)
        for row_idx in range(3, len(template_rows) + 3):
            ws[f"{col_letter}{row_idx}"].number_format = "@"
    wb.save(output_path)


def process_single(zip3, output_path, index=None):
    if index is None:
        index = get_zone_index()

    xls_url = find_xls_url(index, zip3)
    if not xls_url:
        raise RuntimeError(f"ZIP3 {zip3} not found in index")

    with tempfile.NamedTemporaryFile(suffix=".xls", delete=False) as tmp:
        tmp_xls = tmp.name

    try:
        for attempt in range(MAX_RETRIES):
            try:
                size = download_xls(xls_url, tmp_xls)
                break
            except Exception as e:
                if attempt < MAX_RETRIES - 1:
                    delay = RETRY_DELAYS[attempt]
                    print(f"  Retry {attempt + 1}/{MAX_RETRIES} after {delay}s: {e}", file=sys.stderr)
                    time.sleep(delay)
                else:
                    raise

        if size < 100:
            raise RuntimeError(f"Downloaded file too small ({size} bytes)")

        main_rows, footnote_rows = parse_xls_ground(tmp_xls)
        template = build_template_rows(main_rows, footnote_rows)
        write_excel(template, output_path)
        return len(template)
    finally:
        if os.path.exists(tmp_xls):
            os.unlink(tmp_xls)


def load_checkpoint(checkpoint_path):
    if os.path.exists(checkpoint_path):
        with open(checkpoint_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"completed": [], "failed": []}


def save_checkpoint(checkpoint_path, data):
    tmp = checkpoint_path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, checkpoint_path)


def batch_download(output_dir, checkpoint_path=None):
    if checkpoint_path is None:
        checkpoint_path = os.path.join(output_dir, "checkpoint.json")

    os.makedirs(output_dir, exist_ok=True)
    checkpoint = load_checkpoint(checkpoint_path)
    completed = set(checkpoint.get("completed", []))
    failed_list = list(checkpoint.get("failed", []))

    print(f"Fetching zone index...", file=sys.stderr)
    index = get_zone_index()
    total = len(index["data"])
    print(f"Index: {total} ZIP3 entries, {len(completed)} already done", file=sys.stderr)

    new_failed = []
    for i, entry in enumerate(index["data"]):
        zip3 = entry["zip"]
        if zip3 in completed:
            continue

        xlsx_name = f"UPS-GROUND-{zip3}.xlsx"
        xlsx_path = os.path.join(output_dir, xlsx_name)

        progress = len(completed) + len(new_failed) + 1
        print(f"[{progress}/{total}] Processing ZIP3={zip3}...", end=" ", file=sys.stderr)

        try:
            row_count = process_single(zip3, xlsx_path, index=index)
            completed.add(zip3)
            checkpoint["completed"] = sorted(completed)
            save_checkpoint(checkpoint_path, checkpoint)
            print(f"OK ({row_count} rows)", file=sys.stderr)
        except Exception as e:
            new_failed.append({"zip3": zip3, "error": str(e)})
            print(f"FAILED: {e}", file=sys.stderr)

        time.sleep(0.5)

    if new_failed:
        all_failed = failed_list + new_failed
        checkpoint["failed"] = all_failed
        save_checkpoint(checkpoint_path, checkpoint)

    print(f"\nDone: {len(completed)} success, {len(new_failed)} failed", file=sys.stderr)
    if new_failed:
        for f in new_failed:
            print(f"  FAILED: {f['zip3']}: {f['error']}", file=sys.stderr)

    return {
        "total": total,
        "completed": len(completed),
        "failed": len(new_failed),
        "output_dir": output_dir,
    }


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Download UPS zone charts and convert to template Excel")
    parser.add_argument("--zip3", default="", help="Single ZIP3 to download (e.g. 910)")
    parser.add_argument("--all", action="store_true", help="Download all 902 ZIP3 prefixes")
    parser.add_argument("--output", default="", help="Output path (single mode) or output directory (batch mode)")
    parser.add_argument("--checkpoint", default="", help="Checkpoint file path (batch mode, default: <output_dir>/checkpoint.json)")
    args = parser.parse_args()

    if not args.zip3 and not args.all:
        parser.error("Specify --zip3 <code> or --all")

    try:
        if args.zip3:
            output = args.output or os.path.join(
                os.path.expanduser("~"), "Desktop", f"UPS-GROUND-{args.zip3}.xlsx"
            )
            row_count = process_single(args.zip3, output)
            print(json.dumps({
                "status": "success",
                "zip3": args.zip3,
                "output": os.path.abspath(output),
                "rows": row_count,
            }, ensure_ascii=False))
        elif args.all:
            output_dir = args.output or os.path.join(
                os.path.expanduser("~"), ".hermes", "pack-data", "meizheng", "zone-charts"
            )
            result = batch_download(output_dir, args.checkpoint or None)
            print(json.dumps(result, ensure_ascii=False))
    except Exception as e:
        print(json.dumps({"status": "failed", "reason": str(e)}, ensure_ascii=False))
        sys.exit(1)
