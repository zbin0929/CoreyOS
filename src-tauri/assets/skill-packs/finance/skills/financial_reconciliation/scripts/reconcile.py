#!/usr/bin/env python3
"""
Financial Reconciliation Tool — Excel对账引擎

Modes:
  preview  — 预览Excel文件结构和数据
  compare  — 两表互核（按关键列匹配，比对数值列）
  validate — 单表校验（合计、重复、缺失、异常值）
  match    — 模糊匹配（金额+日期近似匹配）

Dependencies auto-installed on first run: pandas, openpyxl
"""

import subprocess
import sys
import os
import argparse
import re
from datetime import datetime, timedelta


def ensure_deps():
    required = {"pandas": "pandas", "openpyxl": "openpyxl"}
    missing = []
    for module, package in required.items():
        try:
            __import__(module)
        except ImportError:
            missing.append(package)
    if missing:
        print(f"[reconcile] Installing: {', '.join(missing)} ...")
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "--quiet"] + missing
        )
        print("[reconcile] Dependencies installed.")


ensure_deps()

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
GREEN_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

LARGE_FILE_THRESHOLD = 100_000
CHUNK_SIZE = 50_000


def clean_numeric(series):
    if series.dtype == object:
        cleaned = series.astype(str).str.replace(",", "", regex=False)
        cleaned = cleaned.str.replace("¥", "", regex=False)
        cleaned = cleaned.str.replace("$", "", regex=False)
        cleaned = cleaned.str.replace("￥", "", regex=False)
        cleaned = cleaned.str.replace(" ", "", regex=False)
        return pd.to_numeric(cleaned, errors="coerce")
    return pd.to_numeric(series, errors="coerce")


def parse_date(series):
    return pd.to_datetime(series, errors="coerce")


def read_excel_smart(path, sheet_name=0):
    file_size = os.path.getsize(path)
    if file_size > 50 * 1024 * 1024:
        print(f"[reconcile] Large file ({file_size // 1024 // 1024}MB), reading...")
    df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
    print(f"[reconcile] Loaded: {len(df)} rows x {len(df.columns)} columns")
    return df


def style_header(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                cell_len = len(str(cell.value or ""))
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 50)


def split_cols(s):
    if not s:
        return []
    return [c.strip() for c in s.split(",") if c.strip()]


# ─────────────────── preview ───────────────────────────────────

def cmd_preview(args):
    path = args.file
    print(f"\n{'='*60}")
    print(f"  File: {os.path.basename(path)}")
    print(f"  Size: {os.path.getsize(path) / 1024:.1f} KB")
    print(f"{'='*60}\n")

    xls = pd.ExcelFile(path, engine="openpyxl")
    for sheet in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet, nrows=5, engine="openpyxl")
        total = pd.read_excel(xls, sheet_name=sheet, engine="openpyxl")
        print(f"--- Sheet: {sheet} ({len(total)} rows) ---")
        print(f"Columns: {list(df.columns)}")
        print(df.to_string(index=False))
        print()


# ─────────────────── compare ──────────────────────────────────

def cmd_compare(args):
    keys = split_cols(args.keys)
    compare_cols = split_cols(args.compare)
    tolerance = args.tolerance
    output = args.output

    if not keys:
        print("[ERROR] --keys is required. Specify the matching column(s).")
        sys.exit(1)

    print(f"[reconcile] Mode: compare")
    print(f"[reconcile] Keys: {keys}")
    print(f"[reconcile] Compare: {compare_cols or 'all non-key columns'}")
    print(f"[reconcile] Tolerance: {tolerance}")

    df_a = read_excel_smart(args.file_a, sheet_name=args.sheet_a or 0)
    df_b = read_excel_smart(args.file_b, sheet_name=args.sheet_b or 0)

    label_a = args.label_a or os.path.basename(args.file_a)
    label_b = args.label_b or os.path.basename(args.file_b)

    for k in keys:
        if k not in df_a.columns:
            print(f"[ERROR] Key column '{k}' not found in file A. Available: {list(df_a.columns)}")
            sys.exit(1)
        if k not in df_b.columns:
            print(f"[ERROR] Key column '{k}' not found in file B. Available: {list(df_b.columns)}")
            sys.exit(1)

    for k in keys:
        df_a[k] = df_a[k].astype(str).str.strip()
        df_b[k] = df_b[k].astype(str).str.strip()

    if not compare_cols:
        common = [c for c in df_a.columns if c in df_b.columns and c not in keys]
        compare_cols = common

    merged = df_a.merge(
        df_b,
        on=keys,
        how="outer",
        indicator=True,
        suffixes=(f"_{label_a}", f"_{label_b}"),
    )

    only_a = merged[merged["_merge"] == "left_only"].copy()
    only_b = merged[merged["_merge"] == "right_only"].copy()
    both = merged[merged["_merge"] == "both"].copy()

    diff_rows = []
    diff_details = []

    for idx, row in both.iterrows():
        row_diffs = {}
        for col in compare_cols:
            col_a = f"{col}_{label_a}"
            col_b = f"{col}_{label_b}"
            if col_a not in both.columns or col_b not in both.columns:
                if col in both.columns:
                    continue
                continue
            val_a = row.get(col_a, None)
            val_b = row.get(col_b, None)

            if pd.isna(val_a) and pd.isna(val_b):
                continue

            try:
                num_a = float(val_a) if not pd.isna(val_a) else None
                num_b = float(val_b) if not pd.isna(val_b) else None
                if num_a is not None and num_b is not None:
                    if abs(num_a - num_b) <= tolerance:
                        continue
                    row_diffs[col] = (val_a, val_b, num_a - num_b)
                    continue
            except (ValueError, TypeError):
                pass

            if str(val_a).strip() != str(val_b).strip():
                row_diffs[col] = (val_a, val_b, None)

        if row_diffs:
            diff_entry = {k: row[k] for k in keys}
            for col, (va, vb, delta) in row_diffs.items():
                diff_entry[f"{col}_{label_a}"] = va
                diff_entry[f"{col}_{label_b}"] = vb
                if delta is not None:
                    diff_entry[f"{col}_差异"] = round(delta, 4)
            diff_rows.append(diff_entry)

    df_diff = pd.DataFrame(diff_rows) if diff_rows else pd.DataFrame()

    summary = {
        "项目": [
            f"{label_a} 总行数",
            f"{label_b} 总行数",
            "匹配行数",
            "有差异行数",
            f"仅在 {label_a} 中",
            f"仅在 {label_b} 中",
            "匹配关键列",
            "比对列",
            "数值容差",
            "生成时间",
        ],
        "数值": [
            len(df_a),
            len(df_b),
            len(both),
            len(df_diff),
            len(only_a),
            len(only_b),
            ", ".join(keys),
            ", ".join(compare_cols),
            tolerance,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ],
    }
    df_summary = pd.DataFrame(summary)

    only_a_clean = only_a.drop(columns=["_merge"], errors="ignore")
    only_b_clean = only_b.drop(columns=["_merge"], errors="ignore")

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="汇总", index=False)
        if not df_diff.empty:
            df_diff.to_excel(writer, sheet_name="差异明细", index=False)
        if not only_a_clean.empty:
            only_a_clean.to_excel(writer, sheet_name=f"仅在{label_a}", index=False)
        if not only_b_clean.empty:
            only_b_clean.to_excel(writer, sheet_name=f"仅在{label_b}", index=False)

        wb = writer.book
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            style_header(ws)
            auto_width(ws)

        if "差异明细" in wb.sheetnames:
            ws = wb["差异明细"]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    col_name = ws.cell(row=1, column=cell.column).value or ""
                    if col_name.endswith("_差异"):
                        if cell.value is not None:
                            try:
                                if float(cell.value) != 0:
                                    cell.fill = RED_FILL
                            except (ValueError, TypeError):
                                pass

    match_rate = (len(both) / max(len(df_a), len(df_b)) * 100) if max(len(df_a), len(df_b)) > 0 else 0

    print(f"\n{'='*60}")
    print(f"  对账完成")
    print(f"{'='*60}")
    print(f"  {label_a}: {len(df_a)} 行")
    print(f"  {label_b}: {len(df_b)} 行")
    print(f"  匹配: {len(both)} 行 ({match_rate:.1f}%)")
    print(f"  差异: {len(df_diff)} 行")
    print(f"  仅在 {label_a}: {len(only_a)} 行")
    print(f"  仅在 {label_b}: {len(only_b)} 行")
    print(f"  报告: {os.path.abspath(output)}")
    print(f"{'='*60}\n")


# ─────────────────── validate ─────────────────────────────────

def cmd_validate(args):
    checks = split_cols(args.checks) if args.checks else ["sum", "duplicates", "missing"]
    output = args.output

    print(f"[reconcile] Mode: validate")
    print(f"[reconcile] Checks: {checks}")

    df = read_excel_smart(args.file, sheet_name=args.sheet or 0)

    results = {}

    if "sum" in checks and args.sum_col:
        sum_cols = split_cols(args.sum_col)
        sum_issues = []
        for col in sum_cols:
            if col not in df.columns:
                sum_issues.append({"列": col, "问题": "列不存在"})
                continue
            numeric = clean_numeric(df[col])
            total = numeric.sum()
            non_null = numeric.count()
            sum_issues.append({
                "列": col,
                "合计": round(total, 2),
                "有效行数": non_null,
                "空值数": len(df) - non_null,
                "最小值": round(numeric.min(), 2) if non_null > 0 else None,
                "最大值": round(numeric.max(), 2) if non_null > 0 else None,
                "平均值": round(numeric.mean(), 2) if non_null > 0 else None,
            })
        results["合计校验"] = pd.DataFrame(sum_issues)

    if "duplicates" in checks:
        key_col = args.key_col
        if key_col:
            dup_cols = split_cols(key_col)
            existing = [c for c in dup_cols if c in df.columns]
            if existing:
                dups = df[df.duplicated(subset=existing, keep=False)]
                if not dups.empty:
                    dups_sorted = dups.sort_values(by=existing)
                    results["重复记录"] = dups_sorted
                    print(f"[reconcile] Found {len(dups)} duplicate rows by {existing}")
                else:
                    results["重复记录"] = pd.DataFrame({"结果": ["未发现重复记录"]})
            else:
                results["重复记录"] = pd.DataFrame({"结果": [f"指定列 {key_col} 不存在"]})
        else:
            dups = df[df.duplicated(keep=False)]
            if not dups.empty:
                results["重复记录"] = dups
            else:
                results["重复记录"] = pd.DataFrame({"结果": ["未发现完全重复行"]})

    if "missing" in checks:
        missing_info = []
        for col in df.columns:
            null_count = df[col].isna().sum()
            empty_count = (df[col].astype(str).str.strip() == "").sum() if df[col].dtype == object else 0
            total_missing = null_count + empty_count
            if total_missing > 0:
                missing_info.append({
                    "列名": col,
                    "空值数": null_count,
                    "空字符串数": empty_count,
                    "缺失率": f"{total_missing / len(df) * 100:.1f}%",
                })
        if missing_info:
            results["缺失值"] = pd.DataFrame(missing_info)
        else:
            results["缺失值"] = pd.DataFrame({"结果": ["所有列均无缺失值"]})

    if "outliers" in checks:
        outlier_rows = []
        numeric_cols = df.select_dtypes(include=["number"]).columns
        for col in numeric_cols:
            series = df[col].dropna()
            if len(series) < 10:
                continue
            mean = series.mean()
            std = series.std()
            if std == 0:
                continue
            z_scores = ((series - mean) / std).abs()
            outlier_idx = z_scores[z_scores > 3].index
            for idx in outlier_idx:
                outlier_rows.append({
                    "行号": idx + 2,
                    "列名": col,
                    "值": df.at[idx, col],
                    "Z-Score": round(z_scores[idx], 2),
                    "均值": round(mean, 2),
                    "标准差": round(std, 2),
                })
        if outlier_rows:
            results["异常值"] = pd.DataFrame(outlier_rows)
        else:
            results["异常值"] = pd.DataFrame({"结果": ["未发现异常值 (Z-Score > 3)"]})

    if "balance" in checks and args.debit_col and args.credit_col:
        debit = clean_numeric(df[args.debit_col]).fillna(0)
        credit = clean_numeric(df[args.credit_col]).fillna(0)
        total_debit = debit.sum()
        total_credit = credit.sum()
        diff = total_debit - total_credit
        balance_info = {
            "项目": ["借方合计", "贷方合计", "差额", "是否平衡"],
            "数值": [
                round(total_debit, 2),
                round(total_credit, 2),
                round(diff, 2),
                "是" if abs(diff) < 0.01 else "否",
            ],
        }
        results["借贷平衡"] = pd.DataFrame(balance_info)

    summary_data = {
        "检查项": [],
        "结果": [],
        "详情": [],
    }
    for name, df_result in results.items():
        if "结果" in df_result.columns and len(df_result) == 1:
            summary_data["检查项"].append(name)
            summary_data["结果"].append("通过")
            summary_data["详情"].append(df_result.iloc[0]["结果"])
        else:
            summary_data["检查项"].append(name)
            summary_data["结果"].append(f"发现 {len(df_result)} 条")
            summary_data["详情"].append(f"详见 [{name}] sheet")
    summary_data["检查项"].append("总行数")
    summary_data["结果"].append(str(len(df)))
    summary_data["详情"].append(f"{len(df.columns)} 列")
    summary_data["检查项"].append("生成时间")
    summary_data["结果"].append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    summary_data["详情"].append("")

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(summary_data).to_excel(writer, sheet_name="汇总", index=False)
        for name, df_result in results.items():
            safe_name = name[:31]
            df_result.to_excel(writer, sheet_name=safe_name, index=False)

        wb = writer.book
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            style_header(ws)
            auto_width(ws)

    print(f"\n{'='*60}")
    print(f"  校验完成")
    print(f"{'='*60}")
    for i, check in enumerate(summary_data["检查项"]):
        print(f"  {check}: {summary_data['结果'][i]}")
    print(f"  报告: {os.path.abspath(output)}")
    print(f"{'='*60}\n")


# ─────────────────── match ────────────────────────────────────

def cmd_match(args):
    tolerance = args.tolerance
    date_range = args.date_range
    output = args.output

    print(f"[reconcile] Mode: fuzzy match")
    print(f"[reconcile] Amount tolerance: {tolerance}")
    print(f"[reconcile] Date range: {date_range} days")

    df_a = read_excel_smart(args.file_a, sheet_name=args.sheet_a or 0)
    df_b = read_excel_smart(args.file_b, sheet_name=args.sheet_b or 0)

    label_a = args.label_a or os.path.basename(args.file_a)
    label_b = args.label_b or os.path.basename(args.file_b)

    amt_a_col = args.amount_a
    amt_b_col = args.amount_b
    date_a_col = args.date_a
    date_b_col = args.date_b

    for col, name, df_ in [
        (amt_a_col, "amount_a", df_a),
        (amt_b_col, "amount_b", df_b),
    ]:
        if col not in df_.columns:
            print(f"[ERROR] Column '{col}' not found. Available: {list(df_.columns)}")
            sys.exit(1)

    df_a["_amt"] = clean_numeric(df_a[amt_a_col])
    df_b["_amt"] = clean_numeric(df_b[amt_b_col])

    use_date = False
    if date_a_col and date_b_col:
        if date_a_col in df_a.columns and date_b_col in df_b.columns:
            df_a["_date"] = parse_date(df_a[date_a_col])
            df_b["_date"] = parse_date(df_b[date_b_col])
            use_date = True

    matched_a = set()
    matched_b = set()
    matches = []

    for i, row_a in df_a.iterrows():
        if i in matched_a:
            continue
        amt_val = row_a["_amt"]
        if pd.isna(amt_val):
            continue

        candidates = df_b[
            (~df_b.index.isin(matched_b))
            & ((df_b["_amt"] - amt_val).abs() <= tolerance)
        ]

        if use_date and not pd.isna(row_a.get("_date")):
            date_val = row_a["_date"]
            candidates = candidates[
                (candidates["_date"] - date_val).abs() <= timedelta(days=date_range)
            ]

        if not candidates.empty:
            best_idx = (candidates["_amt"] - amt_val).abs().idxmin()
            matched_a.add(i)
            matched_b.add(best_idx)

            match_entry = {}
            match_entry[f"{label_a}_行号"] = i + 2
            match_entry[f"{label_b}_行号"] = best_idx + 2
            match_entry[f"{amt_a_col}_{label_a}"] = row_a[amt_a_col]
            match_entry[f"{amt_b_col}_{label_b}"] = df_b.at[best_idx, amt_b_col]
            match_entry["金额差异"] = round(float(amt_val - df_b.at[best_idx, "_amt"]), 4)
            if use_date:
                match_entry[f"{date_a_col}_{label_a}"] = row_a.get(date_a_col)
                match_entry[f"{date_b_col}_{label_b}"] = df_b.at[best_idx, date_b_col]
            matches.append(match_entry)

    df_matched = pd.DataFrame(matches) if matches else pd.DataFrame()
    unmatched_a = df_a[~df_a.index.isin(matched_a)].drop(columns=["_amt", "_date"], errors="ignore")
    unmatched_b = df_b[~df_b.index.isin(matched_b)].drop(columns=["_amt", "_date"], errors="ignore")

    summary = {
        "项目": [
            f"{label_a} 总行数",
            f"{label_b} 总行数",
            "成功匹配",
            f"{label_a} 未匹配",
            f"{label_b} 未匹配",
            "金额容差",
            "日期容差(天)",
            "生成时间",
        ],
        "数值": [
            len(df_a),
            len(df_b),
            len(df_matched),
            len(unmatched_a),
            len(unmatched_b),
            tolerance,
            date_range if use_date else "未使用",
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ],
    }

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(summary).to_excel(writer, sheet_name="汇总", index=False)
        if not df_matched.empty:
            df_matched.to_excel(writer, sheet_name="已匹配", index=False)
        if not unmatched_a.empty:
            unmatched_a.to_excel(writer, sheet_name=f"未匹配_{label_a}"[:31], index=False)
        if not unmatched_b.empty:
            unmatched_b.to_excel(writer, sheet_name=f"未匹配_{label_b}"[:31], index=False)

        wb = writer.book
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            style_header(ws)
            auto_width(ws)

    match_rate_a = len(matched_a) / len(df_a) * 100 if len(df_a) > 0 else 0
    match_rate_b = len(matched_b) / len(df_b) * 100 if len(df_b) > 0 else 0

    print(f"\n{'='*60}")
    print(f"  模糊匹配完成")
    print(f"{'='*60}")
    print(f"  {label_a}: {len(df_a)} 行, 匹配 {len(matched_a)} ({match_rate_a:.1f}%)")
    print(f"  {label_b}: {len(df_b)} 行, 匹配 {len(matched_b)} ({match_rate_b:.1f}%)")
    print(f"  报告: {os.path.abspath(output)}")
    print(f"{'='*60}\n")


# ─────────────────── CLI ──────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Excel Financial Reconciliation Tool",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    sub = parser.add_subparsers(dest="command", help="Mode")

    p_preview = sub.add_parser("preview", help="Preview Excel structure")
    p_preview.add_argument("file", help="Excel file path")

    p_compare = sub.add_parser("compare", help="Compare two Excel files")
    p_compare.add_argument("file_a", help="File A path")
    p_compare.add_argument("file_b", help="File B path")
    p_compare.add_argument("--keys", required=True, help="Key columns (comma-separated)")
    p_compare.add_argument("--compare", default="", help="Columns to compare (comma-separated, default: all common)")
    p_compare.add_argument("--sheet-a", default=None, help="Sheet name for file A")
    p_compare.add_argument("--sheet-b", default=None, help="Sheet name for file B")
    p_compare.add_argument("--label-a", default=None, help="Label for file A")
    p_compare.add_argument("--label-b", default=None, help="Label for file B")
    p_compare.add_argument("--tolerance", type=float, default=0.01, help="Numeric tolerance (default: 0.01)")
    p_compare.add_argument("-o", "--output", default="reconciliation_report.xlsx", help="Output file")

    p_validate = sub.add_parser("validate", help="Validate single Excel file")
    p_validate.add_argument("file", help="Excel file path")
    p_validate.add_argument("--checks", default="sum,duplicates,missing,outliers", help="Checks: sum,duplicates,missing,outliers,balance")
    p_validate.add_argument("--sum-col", default=None, help="Column(s) for sum check")
    p_validate.add_argument("--key-col", default=None, help="Column(s) for duplicate check")
    p_validate.add_argument("--debit-col", default=None, help="Debit column (for balance check)")
    p_validate.add_argument("--credit-col", default=None, help="Credit column (for balance check)")
    p_validate.add_argument("--sheet", default=None, help="Sheet name")
    p_validate.add_argument("-o", "--output", default="validation_report.xlsx", help="Output file")

    p_match = sub.add_parser("match", help="Fuzzy match by amount and date")
    p_match.add_argument("file_a", help="File A path")
    p_match.add_argument("file_b", help="File B path")
    p_match.add_argument("--amount-a", required=True, help="Amount column in file A")
    p_match.add_argument("--amount-b", required=True, help="Amount column in file B")
    p_match.add_argument("--date-a", default=None, help="Date column in file A")
    p_match.add_argument("--date-b", default=None, help="Date column in file B")
    p_match.add_argument("--sheet-a", default=None, help="Sheet name for file A")
    p_match.add_argument("--sheet-b", default=None, help="Sheet name for file B")
    p_match.add_argument("--label-a", default=None, help="Label for file A")
    p_match.add_argument("--label-b", default=None, help="Label for file B")
    p_match.add_argument("--tolerance", type=float, default=0.01, help="Amount tolerance")
    p_match.add_argument("--date-range", type=int, default=3, help="Date range in days")
    p_match.add_argument("-o", "--output", default="match_report.xlsx", help="Output file")

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        sys.exit(1)

    if args.command == "preview":
        cmd_preview(args)
    elif args.command == "compare":
        cmd_compare(args)
    elif args.command == "validate":
        cmd_validate(args)
    elif args.command == "match":
        cmd_match(args)


if __name__ == "__main__":
    main()
