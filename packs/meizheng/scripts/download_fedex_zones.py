#!/usr/bin/env python3
import sys
import os
import re
import json
import time
import argparse
import urllib.request

try:
    import websocket
except ImportError:
    websocket = None

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

try:
    import yaml
except ImportError:
    yaml = None

DOWNLOAD_DIR = os.path.expanduser("~/.hermes/downloads")
HERMES_DIR = os.environ.get("HERMES_DIR", os.path.expanduser("~/.hermes"))
CONFIG_PATH = os.path.join(HERMES_DIR, "pack-data", "meizheng", "config", "zone-config.yaml")
VALID_ZONES = {"2", "3", "4", "5", "6", "7", "8", "9", "17"}


def load_fedex_config():
    if yaml is None:
        return None
    if not os.path.exists(CONFIG_PATH):
        return None
    with open(CONFIG_PATH, "r") as f:
        cfg = yaml.safe_load(f)
    return cfg.get("carriers", {}).get("fedex", {})


def get_fedex_tab_ws(cdp_port):
    tabs = json.loads(
        urllib.request.urlopen(f"http://localhost:{cdp_port}/json/list").read()
    )
    for t in tabs:
        if "fedex.com" in t.get("url", "") and t.get("type") == "page":
            return t["webSocketDebuggerUrl"]
    return None


def cdp_call(ws, method, params=None):
    ws.send(json.dumps({"id": 1, "method": method, "params": params or {}}))
    while True:
        r = json.loads(ws.recv())
        if r.get("id") == 1:
            return r


def eval_js(ws, expr):
    r = cdp_call(
        ws,
        "Runtime.evaluate",
        {"expression": expr, "returnByValue": True, "awaitPromise": True},
    )
    return r.get("result", {}).get("result", {}).get("value")


def fetch_pdf_path(ws, zip_code):
    fetch_js = f"""
    (async function() {{
        const form = new URLSearchParams();
        form.set('method', 'GetZoneLocators');
        form.set('downloadFileName', '');
        form.set('fedexZones', 'yes');
        form.set('serviceType', 'domestic');
        form.set('zoneLocatorZipcode', '{zip_code}');
        form.set('zoneLocatorFormat', 'pdf');
        form.set('ratesByServiceFormat', 'pdf');
        form.set('ratesByServiceType', 'List');

        const resp = await fetch('/ratetools/RateToolsMain.do', {{
            method: 'POST',
            headers: {{'Content-Type': 'application/x-www-form-urlencoded'}},
            body: form.toString(),
            credentials: 'include'
        }});
        const html = await resp.text();
        const match = html.match(/downloadFileName" value="([^"]+)"/);
        return JSON.stringify({{
            status: resp.status,
            pdfPath: match ? match[1] : null,
            htmlLen: html.length
        }});
    }})()
    """
    result = eval_js(ws, fetch_js)
    if not result:
        return None
    data = json.loads(result)
    if data.get("status") != 200 or not data.get("pdfPath"):
        print(
            f"  FedEx POST failed: status={data.get('status')}, pdfPath={data.get('pdfPath')}",
            file=sys.stderr,
        )
        return None
    return data["pdfPath"]


def download_pdf(ws, pdf_path):
    pdf_filename = pdf_path.split("/")[-1]
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    dest = os.path.join(DOWNLOAD_DIR, pdf_filename)

    download_js = f"""
    (async function() {{
        const resp = await fetch('/ratetools/{pdf_path}', {{
            credentials: 'include',
            headers: {{'Accept': 'application/pdf,*/*'}}
        }});
        const buf = await resp.arrayBuffer();
        const bytes = new Uint8Array(buf);
        const arr = Array.from(bytes);
        return JSON.stringify({{
            status: resp.status,
            contentType: resp.headers.get('content-type'),
            size: arr.length
        }});
    }})()
    """
    dl_result = eval_js(ws, download_js)
    if not dl_result:
        return None
    dl_data = json.loads(dl_result)

    if dl_data.get("contentType", "").startswith("text/html"):
        print(f"  Got HTML instead of PDF (size={dl_data.get('size')})", file=sys.stderr)
        return None

    save_js = f"""
    (async function() {{
        const resp = await fetch('/ratetools/{pdf_path}', {{
            credentials: 'include'
        }});
        const blob = await resp.blob();
        const url = URL.createObjectURL(blob);
        const a = document.createElement('a');
        a.href = url;
        a.download = '{pdf_filename}';
        document.body.appendChild(a);
        a.click();
        document.body.removeChild(a);
        URL.revokeObjectURL(url);
        return resp.status;
    }})()
    """
    eval_js(ws, save_js)
    time.sleep(3)

    if os.path.exists(dest):
        return dest

    for f in os.listdir(DOWNLOAD_DIR):
        if f == pdf_filename:
            return os.path.join(DOWNLOAD_DIR, f)

    return None


def parse_pdf_zone_data(pdf_path):
    if pdfplumber is None:
        raise ImportError("pdfplumber is required: pip install pdfplumber")

    pdf = pdfplumber.open(pdf_path)
    rows = []

    for page in pdf.pages:
        text = page.extract_text()
        if not text:
            continue

        in_non_contiguous = False
        for line in text.split("\n"):
            line = line.strip()
            if not line:
                continue
            if "Alaska, Hawaii" in line:
                in_non_contiguous = True
                continue

            if in_non_contiguous:
                for m in re.finditer(
                    r"(\d{5})(?:-(\d{5}))?\s+(\d+|NA|\*)\s+(\d+|NA|\*)",
                    line,
                ):
                    start = m.group(1)
                    end = m.group(2) or start
                    ground_zone = m.group(4)
                    if ground_zone in VALID_ZONES:
                        rows.append((f"Zone{ground_zone}", start, end))
            else:
                entries = re.findall(r"(\d{5})-(\d{5})\s+(\d+|NA|\*)", line)
                for start, end, zone in entries:
                    if zone in VALID_ZONES:
                        rows.append((f"Zone{zone}", start, end))

    pdf.close()

    seen = set()
    deduped = []
    for row in rows:
        key = (row[0], row[1], row[2])
        if key not in seen:
            seen.add(key)
            deduped.append(row)

    deduped.sort(key=lambda r: int(r[1]))
    return deduped


def extract_for_zip3(all_rows, zip3):
    z3_start = int(zip3) * 100
    z3_end = z3_start + 99
    matched = []
    for zone_code, start, end in all_rows:
        s = int(start)
        e = int(end)
        if s > z3_end or e < z3_start:
            continue
        clip_s = f"{max(s, z3_start):05d}"
        clip_e = f"{min(e, z3_end):05d}"
        matched.append((zone_code, clip_s, clip_e))
    return matched


def write_excel(rows, output_path):
    if openpyxl is None:
        raise ImportError("openpyxl is required: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    notice = (
        "1、分区代码、开始邮编、截止邮编均为必填\n"
        "2、邮编仅支持数字、字母\n"
        "3、仅能填写已添加的分区代码\n"
        "4、同一个邮编不能同时存在于2个分区内"
    )
    ws.append([notice, "", ""])
    ws.append(["分区代码", "开始邮编", "截止邮编"])
    for row in rows:
        ws.append(list(row))
    for col_idx in range(2, 4):
        col_letter = get_column_letter(col_idx)
        for row_idx in range(3, len(rows) + 3):
            ws[f"{col_letter}{row_idx}"].number_format = "@"
    wb.save(output_path)


def _fetch_pdf_rows(ws_conn, zip3, pdf_cache):
    zip5 = f"{zip3}00"
    pdf_path = fetch_pdf_path(ws_conn, zip5)
    if not pdf_path:
        raise RuntimeError(f"FedEx returned no PDF for ZIP3 {zip3}")

    if pdf_path not in pdf_cache:
        local_pdf = download_pdf(ws_conn, pdf_path)
        if not local_pdf:
            raise RuntimeError(f"Failed to download PDF: {pdf_path}")
        pdf_cache[pdf_path] = parse_pdf_zone_data(local_pdf)
        if local_pdf.startswith(DOWNLOAD_DIR):
            try:
                os.remove(local_pdf)
            except OSError:
                pass

    return pdf_cache[pdf_path]


def process_single(zip3, output_dir, cdp_port=9222):
    if websocket is None:
        raise ImportError("websocket-client is required: pip install websocket-client")

    ws_url = get_fedex_tab_ws(cdp_port)
    if not ws_url:
        raise RuntimeError(f"No FedEx tab found in AI Browser (CDP port {cdp_port})")

    ws = websocket.create_connection(ws_url, suppress_origin=True)
    try:
        print(f"  ZIP3 {zip3} -> {zip3}00...", end=" ", file=sys.stderr)
        all_rows = _fetch_pdf_rows(ws, zip3, {})
        rows = extract_for_zip3(all_rows, zip3)
        if not rows:
            print("no data", file=sys.stderr)
            return None, 0
        print(f"{len(rows)} zones", file=sys.stderr)
    finally:
        ws.close()

    os.makedirs(output_dir, exist_ok=True)
    xlsx_path = os.path.join(output_dir, f"FEDEX-GROUND-{zip3}.xlsx")
    write_excel(rows, xlsx_path)
    return xlsx_path, len(rows)


def batch_download(output_dir, cdp_port=9222, checkpoint_path=None):
    if checkpoint_path is None:
        checkpoint_path = os.path.join(output_dir, "checkpoint.json")

    os.makedirs(output_dir, exist_ok=True)
    cp = _load_checkpoint(checkpoint_path)
    completed = set(cp.get("completed", []))
    failed_list = cp.get("failed", [])

    if websocket is None:
        raise ImportError("websocket-client is required: pip install websocket-client")

    ws_url = get_fedex_tab_ws(cdp_port)
    if not ws_url:
        raise RuntimeError(f"No FedEx tab found in AI Browser (CDP port {cdp_port})")

    ws = websocket.create_connection(ws_url, suppress_origin=True)
    pdf_cache = {}
    try:
        zip3_list = [str(i).zfill(3) for i in range(1000)]
        total = len(zip3_list)
        new_failed = []

        for i, zip3 in enumerate(zip3_list):
            if zip3 in completed:
                continue

            done = len(completed) + len(new_failed) + len(failed_list) + 1
            print(f"[{done}/{total}] {zip3} -> {zip3}00...", end=" ", file=sys.stderr)

            try:
                all_rows = _fetch_pdf_rows(ws, zip3, pdf_cache)
                rows = extract_for_zip3(all_rows, zip3)
                if not rows:
                    completed.add(zip3)
                    cp["completed"] = sorted(completed)
                    _save_checkpoint(checkpoint_path, cp)
                    print("empty (skipped)", file=sys.stderr)
                    continue

                xlsx_path = os.path.join(output_dir, f"FEDEX-GROUND-{zip3}.xlsx")
                write_excel(rows, xlsx_path)
                completed.add(zip3)
                cp["completed"] = sorted(completed)
                _save_checkpoint(checkpoint_path, cp)
                print(f"OK ({len(rows)})", file=sys.stderr)
            except Exception as e:
                new_failed.append({"zip3": zip3, "error": str(e)})
                print(f"FAIL: {e}", file=sys.stderr)

            time.sleep(0.5)
    finally:
        ws.close()

    if new_failed:
        cp["failed"] = failed_list + new_failed
        _save_checkpoint(checkpoint_path, cp)

    print(f"\nDone: {len(completed)} OK, {len(new_failed)} failed", file=sys.stderr)
    return {"total": total, "completed": len(completed), "failed": len(new_failed)}


def _load_checkpoint(path):
    if os.path.exists(path):
        with open(path, "r") as f:
            return json.load(f)
    return {"completed": [], "failed": []}


def _save_checkpoint(path, data):
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Download FedEx zone charts via AI Browser CDP and convert to template Excel"
    )
    parser.add_argument("--zip3", default="", help="3-digit ZIP code (e.g. 910)")
    parser.add_argument("--all", action="store_true", help="Download all ZIP3 (000-999)")
    parser.add_argument("--output-dir", default=None, help="Output directory for Excel files")
    parser.add_argument("--cdp-port", type=int, default=None, help="CDP port (overrides config)")
    args = parser.parse_args()

    if not args.zip3 and not args.all:
        parser.error("Specify --zip3 <code> or --all")

    cfg = load_fedex_config()
    cdp_port = args.cdp_port or (cfg.get("source", {}).get("cdp_port", 9222) if cfg else 9222)
    default_output = os.path.join(HERMES_DIR, "pack-data", "meizheng", "zone-charts", "fedex")
    output_dir = args.output_dir or default_output

    try:
        if args.zip3:
            xlsx_path, row_count = process_single(args.zip3, output_dir, cdp_port)
            if xlsx_path:
                print(json.dumps({
                    "status": "success",
                    "zip3": args.zip3,
                    "output": os.path.abspath(xlsx_path),
                    "rows": row_count,
                }, ensure_ascii=False))
            else:
                print(json.dumps({"status": "empty", "zip3": args.zip3}, ensure_ascii=False))
        elif args.all:
            result = batch_download(output_dir, cdp_port)
            print(json.dumps({"status": "batch_done", **result}, ensure_ascii=False))
    except Exception as e:
        print(json.dumps({"status": "failed", "reason": str(e)}, ensure_ascii=False))
        sys.exit(1)
